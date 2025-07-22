from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import os
import io
import json
from datetime import datetime
import uuid

from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
import openai
from openai import OpenAI

import PyPDF2
from docx import Document
import openpyxl
import markdown
import requests

indexes_storage = {}
documents_storage = {}

openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY", ""))

class CustomQdrantClient:
    """Custom Qdrant client that uses HTTP requests instead of the Qdrant client library"""
    
    def __init__(self, url: str, api_key: Optional[str] = None):
        self.base_url = url.rstrip('/')
        self.headers = {'Content-Type': 'application/json'}
        if api_key:
            self.headers['api-key'] = api_key
    
    def get_collections(self):
        """Get all collections"""
        response = requests.get(f"{self.base_url}/collections", headers=self.headers, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        # Create a mock collections object that matches the Qdrant client interface
        class MockCollections:
            def __init__(self, collections_data):
                self.collections = []
                for col in collections_data.get('result', {}).get('collections', []):
                    class MockCollection:
                        def __init__(self, name):
                            self.name = name
                    self.collections.append(MockCollection(col['name']))
        
        return MockCollections(data)
    
    def get_collection(self, collection_name: str):
        """Get a specific collection"""
        response = requests.get(f"{self.base_url}/collections/{collection_name}", headers=self.headers, timeout=10)
        if response.status_code == 404:
            raise Exception(f"Collection {collection_name} not found")
        response.raise_for_status()
        return response.json()
    
    def create_collection(self, collection_name: str, vectors_config):
        """Create a new collection"""
        payload = {
            "vectors": {
                "size": vectors_config.size,
                "distance": vectors_config.distance.value
            }
        }
        response = requests.put(f"{self.base_url}/collections/{collection_name}", 
                              headers=self.headers, json=payload, timeout=10)
        response.raise_for_status()
        return response.json()
    
    def upsert(self, collection_name: str, points):
        """Upsert points to a collection"""
        payload = {
            "points": [
                {
                    "id": point.id,
                    "vector": point.vector,
                    "payload": point.payload
                } for point in points
            ]
        }
        response = requests.put(f"{self.base_url}/collections/{collection_name}/points", 
                              headers=self.headers, json=payload, timeout=30)
        response.raise_for_status()
        return response.json()
    
    def search(self, collection_name: str, query_vector, limit=10):
        """Search for similar vectors"""
        payload = {
            "vector": query_vector,
            "limit": limit
        }
        response = requests.post(f"{self.base_url}/collections/{collection_name}/points/search", 
                               headers=self.headers, json=payload, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        # Create mock search results that match the Qdrant client interface
        class MockSearchResult:
            def __init__(self, result_data):
                self.id = result_data['id']
                self.score = result_data['score']
                self.payload = result_data.get('payload', {})
        
        return [MockSearchResult(result) for result in data.get('result', [])]
    
    def delete_collection(self, collection_name: str):
        """Delete a collection"""
        response = requests.delete(f"{self.base_url}/collections/{collection_name}", 
                                 headers=self.headers, timeout=10)
        response.raise_for_status()
        return response.json()

def get_qdrant_client(url: str, api_key: Optional[str] = None):
    """Create Qdrant client with provided connection parameters"""
    try:
        # Ensure URL has the correct format
        if not url.startswith(('http://', 'https://')):
            url = f"https://{url}"
        
        # Remove any trailing slashes
        url = url.rstrip('/')
        
        # Test the connection using HTTP requests (like curl)
        headers = {'Content-Type': 'application/json'}
        if api_key:
            headers['api-key'] = api_key
        
        # Test the collections endpoint directly
        test_url = f"{url}/collections"
        
        response = requests.get(test_url, headers=headers, timeout=10)
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail=f"HTTP connection failed: {response.text}")
        
        # Use our custom client instead of the problematic Qdrant client
        client = CustomQdrantClient(url, api_key)
        
        # Test the connection by getting collections
        collections = client.get_collections()
        
        return client
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to connect to Qdrant: {str(e)}")

class DocumentMetadata(BaseModel):
    index_name: str
    description: Optional[str] = ""
    chunk_size: int = 1000
    chunk_overlap: int = 200
    chunking_method: str = "recursive"

class IndexInfo(BaseModel):
    name: str
    description: str
    document_count: int
    created_at: str

class DocumentInfo(BaseModel):
    id: str
    filename: str
    file_type: str
    size: int
    chunks_count: int
    uploaded_at: str

def extract_text_from_file(file_content: bytes, filename: str) -> str:
    """Extract text from various file types"""
    file_extension = filename.lower().split('.')[-1]
    
    try:
        if file_extension == 'pdf':
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
        
        elif file_extension == 'docx':
            doc = Document(io.BytesIO(file_content))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        
        elif file_extension in ['xlsx', 'xls']:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return text
        
        elif file_extension == 'md':
            text_content = file_content.decode('utf-8')
            html = markdown.markdown(text_content)
            import re
            text = re.sub('<[^<]+?>', '', html)
            return text
        
        elif file_extension == 'txt':
            return file_content.decode('utf-8')
        
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_extension}")
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

def chunk_text(text: str, chunk_size: int, chunk_overlap: int, method: str = "recursive") -> List[str]:
    """Chunk text into smaller pieces"""
    if method == "recursive":
        chunks = []
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunk = text[start:end]
            chunks.append(chunk)
            start = end - chunk_overlap
            if start >= len(text):
                break
        return chunks
    else:
        sentences = text.split('. ')
        chunks = []
        current_chunk = ""
        
        for sentence in sentences:
            if len(current_chunk + sentence) < chunk_size:
                current_chunk += sentence + ". "
            else:
                if current_chunk:
                    chunks.append(current_chunk.strip())
                current_chunk = sentence + ". "
        
        if current_chunk:
            chunks.append(current_chunk.strip())
        
        return chunks

async def get_embeddings(texts: List[str]) -> List[List[float]]:
    """Get embeddings from OpenAI or use mock embeddings for testing"""
    try:
        if os.getenv("OPENAI_API_KEY"):
            response = openai_client.embeddings.create(
                model="text-embedding-3-small",
                input=texts
            )
            return [embedding.embedding for embedding in response.data]
        else:
            import random
            import hashlib
            mock_embeddings = []
            for text in texts:
                text_hash = hashlib.md5(text.encode()).hexdigest()
                random.seed(int(text_hash[:8], 16))
                embedding = [random.uniform(-1, 1) for _ in range(1536)]
                mock_embeddings.append(embedding)
            return mock_embeddings
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting embeddings: {str(e)}")

def get_indexes():
    """Get all existing indexes with document counts"""
    result = []
    for index_name, index_data in indexes_storage.items():
        result.append(IndexInfo(
            name=index_name,
            description=index_data.get("description", ""),
            document_count=len(index_data.get("documents", [])),
            created_at=index_data.get("created_at", "")
        ))
    return result

def get_index_documents(index_name: str):
    """Get all documents in a specific index"""
    if index_name not in indexes_storage:
        raise HTTPException(status_code=404, detail="Index not found")
    
    documents = []
    for doc_id in indexes_storage[index_name].get("documents", []):
        if doc_id in documents_storage:
            doc_data = documents_storage[doc_id]
            documents.append(DocumentInfo(
                id=doc_id,
                filename=doc_data["filename"],
                file_type=doc_data["file_type"],
                size=doc_data["size"],
                chunks_count=doc_data["chunks_count"],
                uploaded_at=doc_data["uploaded_at"]
            ))
    return documents

async def upload_document(file: UploadFile, metadata: str, qdrant_url: str, qdrant_api_key: Optional[str] = None):
    """Upload and process a document"""
    try:
        metadata_obj = DocumentMetadata.parse_raw(metadata)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid metadata: {str(e)}")
    
    file_content = await file.read()
    if len(file_content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size exceeds 20MB limit")
    
    text = extract_text_from_file(file_content, file.filename)
    
    chunks = chunk_text(
        text, 
        metadata_obj.chunk_size, 
        metadata_obj.chunk_overlap, 
        metadata_obj.chunking_method
    )
    
    embeddings = await get_embeddings(chunks)
    
    qdrant_client = get_qdrant_client(qdrant_url, qdrant_api_key)
    collection_name = metadata_obj.index_name
    try:
        qdrant_client.get_collection(collection_name)
    except:
        qdrant_client.create_collection(
            collection_name=collection_name,
            vectors_config=VectorParams(size=1536, distance=Distance.COSINE)
        )
    
    doc_id = str(uuid.uuid4())
    
    points = []
    for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
        point_id = str(uuid.uuid4())
        points.append(PointStruct(
            id=point_id,
            vector=embedding,
            payload={
                "document_id": doc_id,
                "chunk_index": i,
                "text": chunk,
                "filename": file.filename,
                "file_type": file.filename.split('.')[-1].lower()
            }
        ))
    
    qdrant_client.upsert(collection_name=collection_name, points=points)
    
    documents_storage[doc_id] = {
        "filename": file.filename,
        "file_type": file.filename.split('.')[-1].lower(),
        "size": len(file_content),
        "chunks_count": len(chunks),
        "uploaded_at": datetime.now().isoformat(),
        "index_name": metadata_obj.index_name
    }
    
    if metadata_obj.index_name not in indexes_storage:
        indexes_storage[metadata_obj.index_name] = {
            "description": metadata_obj.description,
            "created_at": datetime.now().isoformat(),
            "documents": []
        }
    
    indexes_storage[metadata_obj.index_name]["documents"].append(doc_id)
    
    return {
        "document_id": doc_id,
        "filename": file.filename,
        "chunks_processed": len(chunks),
        "status": "success"
    }

def delete_document(index_name: str, document_id: str, qdrant_url: str, qdrant_api_key: Optional[str] = None):
    """Delete a specific document from an index"""
    if index_name not in indexes_storage:
        raise HTTPException(status_code=404, detail="Index not found")
    
    if document_id not in documents_storage:
        raise HTTPException(status_code=404, detail="Document not found")
    
    try:
        from qdrant_client.models import Filter, FieldCondition, MatchValue
        
        qdrant_client = get_qdrant_client(qdrant_url, qdrant_api_key)
        scroll_result = qdrant_client.scroll(
            collection_name=index_name,
            scroll_filter=Filter(
                must=[
                    FieldCondition(
                        key="document_id",
                        match=MatchValue(value=document_id)
                    )
                ]
            )
        )
        
        point_ids = [point.id for point in scroll_result[0]]
        if point_ids:
            qdrant_client.delete(collection_name=index_name, points_selector=point_ids)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting from vector store: {str(e)}")
    
    del documents_storage[document_id]
    indexes_storage[index_name]["documents"].remove(document_id)
    
    return {"status": "success", "message": "Document deleted successfully"}

def delete_index(index_name: str, qdrant_url: str, qdrant_api_key: Optional[str] = None):
    """Delete an entire index"""
    if index_name not in indexes_storage:
        raise HTTPException(status_code=404, detail="Index not found")
    
    try:
        qdrant_client = get_qdrant_client(qdrant_url, qdrant_api_key)
        qdrant_client.delete_collection(collection_name=index_name)
    except Exception as e:
        pass
    
    for doc_id in indexes_storage[index_name]["documents"]:
        if doc_id in documents_storage:
            del documents_storage[doc_id]
    
    del indexes_storage[index_name]
    
    return {"status": "success", "message": "Index deleted successfully"}

async def search_documents(index_name: str, query: dict, qdrant_url: str, qdrant_api_key: Optional[str] = None):
    """Search documents in an index"""
    if index_name not in indexes_storage:
        raise HTTPException(status_code=404, detail="Index not found")
    
    query_text = query.get("query", "")
    limit = query.get("limit", 10)
    
    query_embeddings = await get_embeddings([query_text])
    query_embedding = query_embeddings[0]
    
    try:
        qdrant_client = get_qdrant_client(qdrant_url, qdrant_api_key)
        search_results = qdrant_client.search(
            collection_name=index_name,
            query_vector=query_embedding,
            limit=limit
        )
        
        results = []
        for result in search_results:
            results.append({
                "id": result.id,
                "score": result.score,
                "text": result.payload.get("text", ""),
                "filename": result.payload.get("filename", ""),
                "document_id": result.payload.get("document_id", "")
            })
        
        return {"results": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Search error: {str(e)}")
